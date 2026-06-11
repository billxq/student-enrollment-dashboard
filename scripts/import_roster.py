from __future__ import annotations

import base64
import json
import os
import sqlite3
import shutil
import sys
from copy import deepcopy
from datetime import datetime, timezone
from hashlib import scrypt
from pathlib import Path

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "dashboard-web" / "data.sealed.json"
SQLITE_FILE = ROOT / "dashboard-web" / "data.sqlite"
ARCHIVE_DIR = ROOT / "archive"
ENV_FILE = ROOT / ".env"

REQUIRED_HEADERS = [
    "学生姓名",
    "学生身份证件类型",
    "学生身份证件号",
    "学生当前状态",
    "学生年级",
    "学生班级名称",
    "学生性别",
    "学生户口省份",
    "学生民族",
    "学生学生类别",
]


def read_passphrase() -> str:
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
      if line.startswith("DATA_PASSPHRASE="):
        return line.split("=", 1)[1].strip().strip('"').strip("'")
    return ""


def decrypt_existing(passphrase: str) -> dict:
    sealed = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    salt = base64.b64decode(sealed["salt"])
    iv = base64.b64decode(sealed["iv"])
    tag = base64.b64decode(sealed["tag"])
    ciphertext = base64.b64decode(sealed["ciphertext"])
    key = scrypt(passphrase.encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32)
    return json.loads(AESGCM(key).decrypt(iv, ciphertext + tag, None))


def seal_data(passphrase: str, value: dict) -> dict:
    salt = os.urandom(16)
    iv = os.urandom(12)
    key = scrypt(passphrase.encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32)
    aesgcm = AESGCM(key)
    ciphertext_with_tag = aesgcm.encrypt(iv, json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8"), None)
    return {
        "version": 1,
        "algorithm": "aes-256-gcm",
        "kdf": "scrypt",
        "salt": base64.b64encode(salt).decode("ascii"),
        "iv": base64.b64encode(iv).decode("ascii"),
        "tag": base64.b64encode(ciphertext_with_tag[-16:]).decode("ascii"),
        "ciphertext": base64.b64encode(ciphertext_with_tag[:-16]).decode("ascii"),
    }


def year_number(year_label: str) -> int:
    import re

    match = re.search(r"(\d{4})", year_label or "")
    return int(match.group(1)) if match else 0


def ensure_sqlite_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS academic_years (
            yearLabel TEXT PRIMARY KEY,
            termLabel TEXT NOT NULL,
            sourceFile TEXT NOT NULL,
            sourceName TEXT NOT NULL,
            mtimeMs INTEGER NOT NULL,
            yearNumber INTEGER NOT NULL,
            importedAt TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            yearLabel TEXT NOT NULL REFERENCES academic_years(yearLabel) ON DELETE CASCADE,
            rowIndex INTEGER NOT NULL,
            name TEXT NOT NULL,
            grade TEXT NOT NULL,
            className TEXT NOT NULL,
            gender TEXT NOT NULL,
            status TEXT NOT NULL,
            province TEXT NOT NULL,
            ethnicity TEXT NOT NULL,
            idType TEXT NOT NULL,
            studentCategory TEXT NOT NULL,
            cityStatus TEXT NOT NULL,
            ethnicityGroup TEXT NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_students_year ON students(yearLabel);
        CREATE INDEX IF NOT EXISTS idx_students_year_sort ON students(yearLabel, rowIndex);
        CREATE INDEX IF NOT EXISTS idx_students_year_grade_class_name ON students(yearLabel, grade, className, name);
        """
    )


def read_sqlite_data() -> dict | None:
    if not SQLITE_FILE.exists():
        return None

    conn = sqlite3.connect(SQLITE_FILE)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        ensure_sqlite_schema(conn)
        year_count = conn.execute("SELECT COUNT(*) AS count FROM academic_years").fetchone()["count"]
        if not year_count:
            return None

        generated = conn.execute("SELECT value FROM meta WHERE key = ?", ("generatedAt",)).fetchone()
        year_rows = conn.execute(
            """
            SELECT yearLabel, termLabel, sourceFile, sourceName, mtimeMs, importedAt
            FROM academic_years
            ORDER BY yearNumber DESC, importedAt DESC, yearLabel DESC
            """
        ).fetchall()
        years = []
        for year in year_rows:
            rows = conn.execute(
                """
                SELECT
                    name, grade, className, gender, status, province, ethnicity,
                    idType, studentCategory, cityStatus, ethnicityGroup
                FROM students
                WHERE yearLabel = ?
                ORDER BY rowIndex ASC, id ASC
                """,
                (year["yearLabel"],),
            ).fetchall()
            years.append(
                {
                    "yearLabel": year["yearLabel"],
                    "termLabel": year["termLabel"],
                    "sourceFile": year["sourceFile"],
                    "sourceName": year["sourceName"],
                    "mtimeMs": int(year["mtimeMs"] or 0),
                    "importedAt": year["importedAt"],
                    "rows": [dict(row) for row in rows],
                }
            )

        return {
            "generatedAt": generated["value"] if generated else "",
            "years": years,
        }
    finally:
        conn.close()


def write_sqlite_data(value: dict) -> None:
    SQLITE_FILE.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(SQLITE_FILE)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        ensure_sqlite_schema(conn)
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("DELETE FROM students")
        conn.execute("DELETE FROM academic_years")
        conn.execute("DELETE FROM meta")
        conn.execute(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            ("generatedAt", str(value.get("generatedAt", ""))),
        )

        insert_year = (
            "INSERT OR REPLACE INTO academic_years "
            "(yearLabel, termLabel, sourceFile, sourceName, mtimeMs, yearNumber, importedAt) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)"
        )
        insert_student = (
            "INSERT INTO students "
            "(yearLabel, rowIndex, name, grade, className, gender, status, province, ethnicity, "
            "idType, studentCategory, cityStatus, ethnicityGroup) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        )

        for year in value.get("years", []):
            year_label = str(year.get("yearLabel", "")).strip()
            if not year_label:
                continue
            term_label = str(year.get("termLabel", year_label)).strip() or year_label
            conn.execute(
                insert_year,
                (
                    year_label,
                    term_label,
                    str(year.get("sourceFile", "")).strip(),
                    str(year.get("sourceName", "")).strip(),
                    int(year.get("mtimeMs", 0) or 0),
                    year_number(year_label),
                    str(value.get("generatedAt", "")),
                ),
            )
            for index, row in enumerate(year.get("rows", [])):
                conn.execute(
                    insert_student,
                    (
                        year_label,
                        index,
                        str(row.get("name", "")).strip(),
                        str(row.get("grade", "")).strip(),
                        str(row.get("className", "")).strip(),
                        str(row.get("gender", "")).strip(),
                        str(row.get("status", "")).strip(),
                        str(row.get("province", "")).strip(),
                        str(row.get("ethnicity", "")).strip(),
                        str(row.get("idType", "")).strip(),
                        str(row.get("studentCategory", "")).strip(),
                        str(row.get("cityStatus", "")).strip(),
                        str(row.get("ethnicityGroup", "")).strip(),
                    ),
                )

        conn.commit()
    except:
        conn.rollback()
        raise
    finally:
        conn.close()


def read_existing_data(passphrase: str) -> dict:
    sqlite_data = read_sqlite_data()
    if sqlite_data is not None:
        return sqlite_data

    if DATA_FILE.exists():
        if not passphrase:
            raise RuntimeError("missing DATA_PASSPHRASE")
        return decrypt_existing(passphrase)

    return {"generatedAt": "", "years": []}


def parse_term_label(term_label: str) -> tuple[str, str]:
    import re

    text = term_label.strip()
    match = re.fullmatch(r"(\d{4})\s*学年度\s*(第一|第二|1|2)\s*学期", text)
    if not match:
        raise ValueError("学年度学期格式应为：2025学年度第一学期 或 2025学年度第二学期")
    year = match.group(1)
    semester_raw = match.group(2)
    semester = "第一学期" if semester_raw in {"第一", "1"} else "第二学期"
    return f"{year}学年度", f"{year}学年度{semester}"


def classify_ethnicity(id_type: str, student_category: str, ethnicity: str) -> str:
    text = f"{id_type} {student_category}".lower()
    if any(token in text for token in ("护照", "passport", "外国", "外籍", "港澳", "台胞")):
      return "外籍"
    if ethnicity == "汉族":
      return "汉族"
    if not ethnicity or ethnicity == "其他":
      return "其他/未明确"
    return "少数民族"


def grade_order(text: str) -> int:
    mapping = {
        "一年级": 1,
        "二年级": 2,
        "三年级": 3,
        "四年级": 4,
        "五年级": 5,
        "六年级": 6,
    }
    for key, value in mapping.items():
        if key in text:
            return value
    return 99


def class_order(text: str) -> int:
    import re

    match = re.search(r"(\d+)", text or "")
    return grade_order(text) * 100 + (int(match.group(1)) if match else 99)


def parse_workbook(source: Path) -> list[dict]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise ValueError("Excel 中没有可读取的数据")

    headers = [str(item or "").strip() for item in values[0]]
    index = {name: i for i, name in enumerate(headers)}
    missing = [name for name in REQUIRED_HEADERS if name not in index]
    if missing:
        raise ValueError(f"缺少表头字段：{', '.join(missing)}")

    rows: list[dict] = []
    for row in values[1:]:
        name = str(row[index["学生姓名"]] or "").strip()
        if not name:
            continue
        id_type = str(row[index["学生身份证件类型"]] or "").strip()
        student_category = str(row[index["学生学生类别"]] or "").strip()
        province = str(row[index["学生户口省份"]] or "").strip()
        ethnicity = str(row[index["学生民族"]] or "").strip()
        item = {
            "name": name,
            "grade": str(row[index["学生年级"]] or "").strip(),
            "className": str(row[index["学生班级名称"]] or "").strip(),
            "gender": str(row[index["学生性别"]] or "").strip(),
            "status": str(row[index["学生当前状态"]] or "").strip(),
            "province": province,
            "ethnicity": ethnicity,
            "idType": id_type,
            "studentCategory": student_category,
            "cityStatus": "本市户籍" if province == "上海市" else "非本市户籍",
            "ethnicityGroup": classify_ethnicity(id_type, student_category, ethnicity),
        }
        rows.append(item)

    rows.sort(key=lambda r: (grade_order(r["grade"]), class_order(r["className"]), r["name"]))
    return rows


def copy_to_archive(source: Path, year_label: str, term_label: str) -> Path:
    archive_year_dir = ARCHIVE_DIR / year_label
    archive_year_dir.mkdir(parents=True, exist_ok=True)
    safe_term = term_label.replace(" ", "_")
    target = archive_year_dir / f"{safe_term}_{source.name}"
    if not target.exists():
        shutil.copy2(source, target)
        return target
    suffix = source.suffix
    stem = source.stem
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    alt_target = archive_year_dir / f"{safe_term}_{stem}_{timestamp}{suffix}"
    shutil.copy2(source, alt_target)
    return alt_target


def main() -> None:
    if len(sys.argv) < 4:
        raise SystemExit("usage: import_roster.py <xlsx_path> <year_label> <term_label>")

    source = Path(sys.argv[1]).resolve()
    year_label = sys.argv[2]
    term_label = sys.argv[3]
    passphrase = read_passphrase()

    existing = read_existing_data(passphrase)
    rows = parse_workbook(source)
    archived_path = copy_to_archive(source, year_label, term_label)

    next_years = [deepcopy(year) for year in existing.get("years", []) if year.get("yearLabel") != year_label]
    next_years.append(
        {
            "yearLabel": year_label,
            "termLabel": term_label,
            "sourceFile": archived_path.as_posix(),
            "sourceName": archived_path.name,
            "mtimeMs": int(source.stat().st_mtime * 1000),
            "rows": rows,
        }
    )
    next_years.sort(key=lambda item: int("".join(ch for ch in item.get("yearLabel", "") if ch.isdigit())[:4] or "0"), reverse=True)

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "years": next_years,
    }
    write_sqlite_data(payload)
    if passphrase:
        sealed = seal_data(passphrase, payload)
        DATA_FILE.write_text(json.dumps(sealed, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    foreign_count = sum(1 for row in rows if row["ethnicityGroup"] == "外籍")
    result = {
        "ok": True,
        "yearLabel": year_label,
        "termLabel": term_label,
        "rowCount": len(rows),
        "foreignCount": foreign_count,
        "archiveFile": archived_path.as_posix(),
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
